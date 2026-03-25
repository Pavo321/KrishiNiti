"""
Fetches real fertilizer price data from the World Bank Commodity Price Data (Pink Sheet).
Source: https://www.worldbank.org/en/research/commodity-markets
Columns used: Urea (urea), DAP (phosphate_dap), Potassium Chloride/MOP (potassium)
No API key required. Public domain data.
"""

import hashlib
import json
import sys
from datetime import date
from pathlib import Path

import httpx
import openpyxl

OUTPUT_DIR = Path("data/raw/fertilizer_prices")
SOURCES_FILE = Path("data/sources.md")

# World Bank Pink Sheet — direct Excel download URL
# The file URL changes monthly. This is the stable redirect page to check for latest:
# https://www.worldbank.org/en/research/commodity-markets → "Download data"
# Confirmed working URL as of March 2026:
PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Pink-Sheet-Monthly.xlsx"
)

# Column names in the Pink Sheet Excel (sheet: "Monthly Prices")
# These are the commodity names as they appear in row 5 of the sheet
COMMODITY_COLUMNS = {
    "Urea, E. Europe, fob bulk": "UREA",
    "DAP, fob US Gulf": "DAP",
    "Potassium Chloride, fob Vancouver": "MOP",
}


def download_pink_sheet() -> tuple[bytes, str]:
    print("Downloading World Bank Pink Sheet...")
    headers = {
        "User-Agent": (
            "KrishiNiti-DataPipeline/1.0 "
            "(Agricultural price forecasting research; "
            "github.com/Pavo321/KrishiNiti)"
        )
    }
    with httpx.Client(headers=headers, follow_redirects=True, timeout=120) as client:
        response = client.get(PINK_SHEET_URL)
        response.raise_for_status()

    content = response.content
    sha256 = hashlib.sha256(content).hexdigest()
    print(f"Downloaded {len(content) / 1024:.1f} KB, SHA256: {sha256[:16]}...")
    return content, sha256


def parse_pink_sheet(content: bytes, file_hash: str) -> list[dict]:
    """
    Parse the World Bank Pink Sheet Excel.
    Sheet "Monthly Prices" structure:
    - Row 1-4: headers and metadata
    - Row 5: commodity names
    - Row 6: units
    - Row 7 onward: monthly data, column A = date (e.g., "Jan-60")
    """
    import io

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Try common sheet names
    sheet = None
    for name in ["Monthly Prices", "monthly prices", "Prices"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break

    if sheet is None:
        print(f"Available sheets: {wb.sheetnames}")
        raise ValueError(
            "Could not find 'Monthly Prices' sheet. "
            "World Bank may have updated the file format. "
            "Check the Excel manually and update COMMODITY_COLUMNS."
        )

    # Find the header row (row with commodity names)
    rows = list(sheet.iter_rows(values_only=True))

    # Find header row by looking for known commodity name
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(rows[:10]):
        for j, cell in enumerate(row):
            if cell and "Urea" in str(cell):
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError(
            "Could not find Urea column in first 10 rows. "
            "Pink Sheet format may have changed. Inspect file manually."
        )

    # Map commodity names to column indices
    header_row = rows[header_row_idx]
    for j, cell in enumerate(header_row):
        for source_name, our_name in COMMODITY_COLUMNS.items():
            if cell and source_name.lower() in str(cell).lower():
                col_map[j] = our_name

    if not col_map:
        print(f"Header row contents: {[str(c)[:30] for c in header_row if c]}")
        raise ValueError(
            "Could not map any commodity columns. "
            "Check COMMODITY_COLUMNS against actual sheet headers."
        )

    print(f"Found columns: {col_map}")

    # Parse data rows
    records = []
    for row in rows[header_row_idx + 2 :]:  # skip header + units row
        date_cell = row[0]
        if not date_cell:
            continue

        # Parse date — format is "Jan-60", "Feb-24", etc.
        try:
            if isinstance(date_cell, str) and "-" in date_cell:
                month_str, year_str = date_cell.strip().split("-")
                year = int(year_str)
                year = year + 2000 if year < 50 else year + 1900
                month_num = {
                    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
                    "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
                    "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
                }[month_str[:3]]
                price_date = date(year, month_num, 1)
            elif hasattr(date_cell, "year"):
                price_date = date_cell.date() if hasattr(date_cell, "date") else date_cell
            else:
                continue
        except (ValueError, KeyError):
            continue

        for col_idx, commodity in col_map.items():
            if col_idx >= len(row):
                continue
            price_val = row[col_idx]
            if price_val is None or price_val == "":
                continue
            try:
                price_usd = float(price_val)
            except (ValueError, TypeError):
                continue

            records.append(
                {
                    "price_date": price_date.isoformat(),
                    "commodity": commodity,
                    "price_usd": price_usd,
                    "unit": "USD_PER_MT",
                    "source": "WORLDBANK",
                    "raw_file_hash": file_hash,
                }
            )

    return records


def save_records(records: list[dict], file_hash: str) -> Path:
    today = date.today().isoformat()
    out_path = OUTPUT_DIR / f"worldbank_pinksheet_{today}.json"
    with open(out_path, "w") as f:
        json.dump(records, f, indent=2, default=str)

    # Save checksum alongside
    checksum_path = OUTPUT_DIR / f"worldbank_pinksheet_{today}.sha256"
    checksum_path.write_text(file_hash)

    print(f"Saved {len(records)} price records to {out_path}")
    return out_path


def update_sources_md(record_count: int, date_range: tuple, file_hash: str) -> None:
    entry = f"""
## Fertilizer Prices — World Bank Commodity Price Data (Pink Sheet)
- **URL**: {PINK_SHEET_URL}
- **What it contains**: Monthly global prices for Urea, DAP, MOP in USD/MT since 1960
- **Format**: Excel (.xlsx)
- **Update frequency**: Monthly
- **Date first accessed**: {date.today().isoformat()}
- **Records**: {record_count} rows, {date_range[0]} to {date_range[1]}
- **SHA256**: {file_hash[:32]}...
- **How to refresh**: Run `python scripts/fetch_worldbank_pinksheet.py`
- **Notes**: Column names verified in 'Monthly Prices' sheet. Prices in USD/MT, convert to INR in seed_db.py using RBI exchange rate.
"""
    with open(SOURCES_FILE, "a") as f:
        f.write(entry)
    print(f"Updated {SOURCES_FILE}")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    content, file_hash = download_pink_sheet()

    # Save raw Excel
    raw_path = OUTPUT_DIR / f"CMO-Pink-Sheet-{date.today().isoformat()}.xlsx"
    raw_path.write_bytes(content)

    records = parse_pink_sheet(content, file_hash)

    if not records:
        print("ERROR: No records parsed. Inspect the Excel file manually.")
        sys.exit(1)

    dates = [r["price_date"] for r in records]
    date_range = (min(dates), max(dates))
    print(f"Parsed {len(records)} records spanning {date_range[0]} to {date_range[1]}")

    # Spot check — verify data looks real
    urea_records = [r for r in records if r["commodity"] == "UREA"]
    if urea_records:
        sample = urea_records[-1]
        print(f"Latest Urea price: ${sample['price_usd']}/MT on {sample['price_date']}")
        if not (50 < sample["price_usd"] < 2000):
            print(
                f"WARNING: Urea price ${sample['price_usd']}/MT seems outside "
                "reasonable range ($50–$2000). Verify manually."
            )

    save_records(records, file_hash)
    update_sources_md(len(records), date_range, file_hash)
    print("Done.")


if __name__ == "__main__":
    main()
