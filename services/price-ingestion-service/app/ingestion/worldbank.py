"""
Fetches latest fertilizer prices from World Bank Pink Sheet.
Called daily by the ingestion job. Idempotent.
"""

import hashlib
import io
import logging
from datetime import date

import httpx
import openpyxl

logger = logging.getLogger(__name__)

PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)

COMMODITY_COLUMNS = {
    "Urea": "UREA",
    "DAP": "DAP",
    "Potassium chloride": "MOP",
    "TSP": "TSP",
}

USD_TO_INR_DEFAULT = 83.5  # Updated by RBI rate fetcher in production


def fetch_latest_prices() -> list[dict]:
    """
    Downloads and parses the World Bank Pink Sheet.
    Returns list of price records for the most recent available month.
    Only returns records not already in the DB (caller checks by date).
    """
    logger.info("Fetching World Bank Pink Sheet...")

    headers = {
        "User-Agent": "KrishiNiti-DataPipeline/1.0 (research; github.com/Pavo321/KrishiNiti)"
    }

    with httpx.Client(headers=headers, follow_redirects=True, timeout=120) as client:
        response = client.get(PINK_SHEET_URL)
        response.raise_for_status()

    content = response.content
    file_hash = hashlib.sha256(content).hexdigest()
    logger.info(f"Downloaded {len(content) / 1024:.1f} KB, hash: {file_hash[:16]}")

    records = _parse_excel(content, file_hash)
    logger.info(f"Parsed {len(records)} price records from Pink Sheet")
    return records


def _parse_excel(content: bytes, file_hash: str) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    sheet = None
    for name in ["Monthly Prices", "monthly prices", "Prices"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break

    if sheet is None:
        raise ValueError(f"Monthly Prices sheet not found. Sheets: {wb.sheetnames}")

    rows = list(sheet.iter_rows(values_only=True))
    header_row_idx, col_map = _find_columns(rows)

    records = []
    for row in rows[header_row_idx + 2:]:
        price_date = _parse_date(row[0])
        if not price_date:
            continue

        for col_idx, commodity in col_map.items():
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            try:
                price_usd = float(row[col_idx])
            except (ValueError, TypeError):
                continue

            records.append({
                "price_date": price_date,
                "commodity": commodity,
                "price_usd": price_usd,
                "price_inr": round(price_usd * USD_TO_INR_DEFAULT * 50 / 1000, 2),
                "unit": "INR_PER_BAG",
                "source": "WORLDBANK",
                "exchange_rate": USD_TO_INR_DEFAULT,
                "raw_file_hash": file_hash,
            })

    return records


def _find_columns(rows: list) -> tuple[int, dict]:
    for i, row in enumerate(rows[:10]):
        col_map = {}
        for j, cell in enumerate(row):
            if not cell:
                continue
            cell_str = str(cell)
            for keyword, commodity in COMMODITY_COLUMNS.items():
                if keyword.lower() in cell_str.lower():
                    col_map[j] = commodity
        if col_map:
            return i, col_map
    raise ValueError("Could not find commodity columns in Pink Sheet header rows.")


def _parse_date(cell) -> date | None:
    if not cell:
        return None
    try:
        cell_str = str(cell).strip()

        # Format: "1960M01" (World Bank current format)
        if "M" in cell_str and len(cell_str) == 7:
            year = int(cell_str[:4])
            month = int(cell_str[5:7])
            return date(year, month, 1)

        # Format: "Jan-60" or "Jan-1960" (legacy)
        if "-" in cell_str:
            parts = cell_str.split("-")
            month_map = {
                "jan": 1, "feb": 2, "mar": 3, "apr": 4,
                "may": 5, "jun": 6, "jul": 7, "aug": 8,
                "sep": 9, "oct": 10, "nov": 11, "dec": 12,
            }
            year = int(parts[1])
            year = year + 2000 if year < 50 else (year + 1900 if year < 100 else year)
            return date(year, month_map[parts[0][:3].lower()], 1)

        # Excel date object
        if hasattr(cell, "year"):
            return cell.date() if hasattr(cell, "date") else cell
    except (ValueError, KeyError, IndexError, AttributeError):
        pass
    return None
